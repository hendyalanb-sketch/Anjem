import openpyxl, json, sys, io, re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('Survei Kebutuhan Rute dan Titik Jemput Pegawai RS Adhyaksa Jatim (Jawaban).xlsx', data_only=True)
ws = wb['Form Responses 1']
headers = [c.value for c in ws[1]]
data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]:
        data.append(dict(zip(headers, row)))

LANDMARKS = [
    (r"rs ?sakinah|rsi ?sakinah|sakinah", "RSI Sakinah Sooko", -7.4878, 112.4284, "Sooko", "mjk-kota"),
    (r"dian ?husada|gemekan|kos ?aulia|dogodogi", "RS Dian Husada / Alfamart Gemekan", -7.5115, 112.4241, "Sooko", "mjk-kota"),
    (r"indomaret veteran mojoagung|mojoagung.*veteran|veteran.*mojoagung|puskesmas mojoagung|miagan", "Indomaret Veteran Mojoagung", -7.5604, 112.3478, "Mojoagung", "mojoagung"),
    (r"smkn ?mojoagung|smk ?mojoagung", "SMKN Mojoagung", -7.5588, 112.3450, "Mojoagung", "mojoagung"),
    (r"mojoagung", "Mojoagung Pusat", -7.5604, 112.3478, "Mojoagung", "mojoagung"),
    (r"indomaret.*brangkal|kelurahan brangkal|jalan raya.*brangkal|jalan raya nrangkal|raya brangkal", "Indomaret Kelurahan Brangkal", -7.5018, 112.4548, "Sooko", "mjk-kota"),
    (r"ahas.*brangkal|pertigaan brangkal", "AHAS Pertigaan Brangkal", -7.5008, 112.4548, "Sooko", "mjk-kota"),
    (r"pom ?brangkal|spbu brangkal", "POM Brangkal", -7.4985, 112.4548, "Sooko", "mjk-kota"),
    (r"terminal kertajaya|terminal\b|kertajaya|kertaja", "Terminal Kertajaya Mojokerto", -7.4759, 112.4470, "Mojokerto Kota", "mjk-kota"),
    (r"dlanggu|smkn ?1 dlanggu|pom bensin dlanggu", "SPBU Dlanggu / SMKN 1 Dlanggu", -7.5829, 112.4980, "Dlanggu", "mojosari"),
    (r"masjid majapahit|majapahit trowulan|darul muttaqin|jatipasar|halaman masjid|mesjid majapahit", "Masjid Majapahit Trowulan", -7.5494, 112.3830, "Trowulan", "trowulan"),
    (r"halte rsk|mojowarno|rsk\.? mojowarno", "Halte RSK Mojowarno", -7.5896, 112.3260, "Jombang", "mojoagung"),
    (r"syech jumadil|jumadil kubro", "Indomaret Syech Jumadil Kubro", -7.5469, 112.3868, "Trowulan", "trowulan"),
    (r"stasiun mojokerto|stasiun curahmalang|stasiun\b", "Stasiun Mojokerto", -7.4720, 112.4350, "Mojokerto Kota", "mjk-kota"),
    (r"klenteng", "Klenteng / Stasiun Mojokerto", -7.4720, 112.4350, "Mojokerto Kota", "mjk-kota"),
    (r"indomaret.*lengkong|lengkong", "Indomaret Lengkong", -7.5780, 112.5550, "Mojosari", "mojosari"),
    (r"waris domas|domas|bidan frina", "Bengkel Las Domas Trowulan", -7.5470, 112.4060, "Trowulan", "trowulan"),
    (r"karangkuten|sd.*karangkuten", "Pertigaan SD Karangkuten Gondang", -7.6080, 112.4750, "Gondang", "mojosari"),
    (r"puskesmas sooko|sooko.*puskesmas", "Puskesmas Sooko", -7.4877, 112.4327, "Sooko", "mjk-kota"),
    (r"smk jatirejo|cc8c\+mxx|sumengko", "SMK Jatirejo (Sumengko)", -7.5928, 112.4135, "Jatirejo", "mjk-kota"),
    (r"toko khayala|sumber jaya|gang rumah", "Sekitar Pungging (Sumber Jaya)", -7.5476, 112.5717, "Pungging", "mojosari"),
    (r"indomaret mlaten|mlaten", "Indomaret Mlaten Puri", -7.5215, 112.4622, "Puri", "puri"),
    (r"plososari|baitul mut|baitul mut.aqiin", "Masjid Baitul Muttaqiin Plososari", -7.5300, 112.4615, "Puri", "puri"),
    (r"polres mojosari|polres kab", "Depan Polres Mojokerto", -7.5260, 112.5404, "Mojosari", "mojosari"),
    (r"spbu badung|pom badung|badung dlanggu", "SPBU Badung Dlanggu", -7.5856, 112.4930, "Dlanggu", "mojosari"),
    (r"gajah mada.*jetis|mlirip jetis|gajah mada mlirip", "Pos Polisi Gajah Mada Mlirip Jetis", -7.4445, 112.4280, "Jetis", "mjk-kota"),
    (r"janti|pacet|pandanarum", "POM Pandanarum / Janti Pacet", -7.6285, 112.5380, "Pacet", "mojosari"),
    (r"canggu|sidorejo surya|indomaret canggu|indomaret jetis|perempatan jetis", "Indomaret Canggu / Sidorejo Jetis", -7.4565, 112.4445, "Jetis", "mjk-kota"),
    (r"pohjejer|gadjah", "Pasar Pohjejer", -7.6090, 112.4710, "Pohjejer", "mojosari"),
    (r"sumberaji|karangjeruk", "Masjid Sumberaji Karangjeruk Jatirejo", -7.6155, 112.4130, "Jatirejo", "mjk-kota"),
    (r"mertex|halte transjat", "Perempatan Mertex Mojoanyar", -7.4815, 112.4685, "Mojoanyar", "mjk-kota"),
    (r"alfamart.*sampangagung|spbu sampangagung|multi bintang", "SPBU Sampangagung (PT Multi Bintang)", -7.5520, 112.5400, "Kutorejo", "mojosari"),
    (r"madasa|masjid agung darussalam|lesehan nikmat|bhayangkara", "MADASA Sooko (Lesehan Nikmat)", -7.4863, 112.4338, "Sooko", "mjk-kota"),
    (r"kpp|pajak mojokerto|kedungpring|ra basuni", "KPP Pratama Mojokerto", -7.4858, 112.4348, "Sooko", "mjk-kota"),
    (r"bpjs ketenagakerjaan|bpjs", "BPJS Ketenagakerjaan Kab. Mojokerto", -7.4970, 112.4541, "Puri", "puri"),
    (r"kantor kecamatan mojosari|kecamatan mojosari", "Kantor Kecamatan Mojosari", -7.5197, 112.5413, "Mojosari", "mojosari"),
    (r"kantor kecamatan sooko|kec.* sooko|kec\.sooko|kecamatan sooko", "Indomaret Kantor Kec. Sooko", -7.4868, 112.4310, "Sooko", "mjk-kota"),
    (r"perempatan sooko|ufo elektronik", "Perempatan Sooko (Ufo Elektronik)", -7.4868, 112.4307, "Sooko", "mjk-kota"),
    (r"alfamart kejari|kejari", "Alfamart Kejari Sooko", -7.4885, 112.4327, "Sooko", "mjk-kota"),
    (r"japan steak", "Japan Steak Sooko", -7.4880, 112.4338, "Sooko", "mjk-kota"),
    (r"man 2|sma man", "MAN 2 Mojokerto Sooko", -7.4790, 112.4280, "Sooko", "mjk-kota"),
    (r"pendopo agung|brawijaya|trowulan.*pendopo|perempatan.*pendopo", "Perempatan Pendopo Agung Trowulan", -7.5476, 112.3848, "Trowulan", "trowulan"),
    (r"makam troloyo|troloyo", "Makam Troloyo Trowulan", -7.5590, 112.3870, "Trowulan", "trowulan"),
    (r"pei hai|peihai|pabrik sepatu", "Pabrik Sepatu Pei Hai Jombang", -7.5510, 112.2330, "Jombang", "mojoagung"),
    (r"rsud ploso|ploso jombang", "RSUD Ploso Jombang", -7.4156, 112.2222, "Plandaan", "mojoagung"),
    (r"pasar peterongan|peterongan", "Pasar Peterongan Jombang", -7.5556, 112.2954, "Kesamben", "mojoagung"),
    (r"warung lumayan|jatipelem|gudo|diwek", "Jatipelem Diwek Jombang", -7.5560, 112.2780, "Gudo", "mojoagung"),
    (r"rsud soekandar|soekandar", "RSUD Soekandar Mojosari", -7.5028, 112.5539, "Mojosari", "mojosari"),
    (r"puskesmas pungging|terminal baru mojosari|pungging.*terminal", "Puskesmas Pungging / Terminal Baru Mojosari", -7.5476, 112.5717, "Pungging", "mojosari"),
    (r"perumahan bumi mojopahit|bumi mojopahit", "Perumahan Bumi Mojopahit Pungging", -7.5276, 112.5800, "Pungging", "mojosari"),
    (r"bungurasih", "Terminal Bungurasih Sidoarjo", -7.3470, 112.7270, "Waru", "mjk-kota"),
    (r"taman makam pahlawan|jl.* pahlawan|pahlawan kota", "Jl. Pahlawan Mojokerto", -7.4708, 112.4327, "Mojokerto Kota", "mjk-kota"),
    (r"masjid rohmat|gading", "Masjid Rohmat Gading Jatirejo", -7.5928, 112.4380, "Jatirejo", "mjk-kota"),
    (r"puskesmas bangsal|honda sekawan|sekawan motor|spn polda", "Puskesmas Bangsal", -7.5304, 112.5046, "Bangsal", "mojosari"),
    (r"dinoyo|pom dinoyo|pom bensin dinoyo", "POM Dinoyo Jatirejo", -7.5630, 112.4470, "Jatirejo", "mjk-kota"),
    (r"kintelan|pertigaan masjid kintelan", "Pertigaan Masjid Kintelan Puri", -7.5215, 112.4640, "Puri", "puri"),
    (r"gapura puri|puri kencana", "Gapura Puri Kencana", -7.5215, 112.4640, "Puri", "puri"),
    (r"sumbertanggu|patung macan", "Patung Macan Sumbertanggu Mojosari", -7.5244, 112.5409, "Mojosari", "mojosari"),
    (r"ketintang", "Masjid Ketintang Gondang", -7.6114, 112.5009, "Gondang", "mojosari"),
    (r"puskesmas jatirejo|jatirejo.*puskesmas", "Puskesmas Jatirejo", -7.6044, 112.4106, "Jatirejo", "mjk-kota"),
    (r"ismul haq|sppg", "SPPG Ismul Haq Jatirejo", -7.5510, 112.4030, "Jatirejo", "trowulan"),
    (r"sdn kumitir|kumitir", "SDN Kumitir Jatirejo", -7.5940, 112.4180, "Jatirejo", "mjk-kota"),
    (r"warung kopi 5r|tol mojokerto", "Exit Tol Mojokerto Jetis", -7.4520, 112.4555, "Jetis", "mjk-kota"),
    (r"toko toni|nurul nur", "Kos Sebelah Toko Toni Puri", -7.5215, 112.4640, "Puri", "puri"),
    (r"miji lama|kranggan|artha prima", "Miji Lama Gg.1 Kranggan", -7.4720, 112.4250, "Mojokerto Kota", "mjk-kota"),
    (r"trowulan", "Trowulan Pusat", -7.5494, 112.3830, "Trowulan", "trowulan"),
    (r"mojosari", "Mojosari Pusat", -7.5197, 112.5413, "Mojosari", "mojosari"),
    (r"jombang", "Jombang Pusat", -7.5510, 112.2330, "Jombang", "mojoagung"),
    (r"pungging", "Pungging Pusat", -7.5476, 112.5717, "Pungging", "mojosari"),
    (r"puri", "Puri Pusat", -7.5215, 112.4622, "Puri", "puri"),
    (r"sooko", "Sooko Pusat", -7.4877, 112.4327, "Sooko", "mjk-kota"),
    (r"mojokerto kota|kota mojokerto", "Mojokerto Kota Pusat", -7.4720, 112.4350, "Mojokerto Kota", "mjk-kota"),
    (r"gondang", "Gondang Pusat", -7.6080, 112.4750, "Gondang", "mojosari"),
    (r"jatirejo", "Jatirejo Pusat", -7.5930, 112.4135, "Jatirejo", "mjk-kota"),
    (r"kutorejo", "Kutorejo Pusat", -7.5520, 112.5400, "Kutorejo", "mojosari"),
    (r"bangsal", "Bangsal Pusat", -7.5304, 112.5046, "Bangsal", "mojosari"),
    (r"ngoro", "Ngoro Pusat", -7.6011, 112.5645, "Ngoro", "mojosari"),
    (r"perak|gadingmangu", "Perak Gadingmangu Jombang", -7.4855, 112.3050, "Perak", "mojoagung"),
    (r"jogoroto", "Jogoroto Jombang", -7.5664, 112.2710, "Jogoroto", "mojoagung"),
    (r"kesamben", "Kesamben Jombang", -7.5570, 112.3380, "Kesamben", "mojoagung"),
    (r"plandaan", "Plandaan Jombang", -7.4156, 112.2222, "Plandaan", "mojoagung"),
    (r"sidoarjo|candi", "Sidoarjo Pusat", -7.4470, 112.7180, "Sidoarjo", "mjk-kota"),
    (r"kebomas|gresik", "Gresik / Kebomas", -7.1545, 112.6360, "Kebomas", "mjk-kota"),
    (r"surabaya|sambikerep", "Surabaya", -7.2575, 112.7521, "Surabaya", "mjk-kota"),
    (r"waru", "Waru Sidoarjo", -7.3470, 112.7270, "Waru", "mjk-kota"),
    (r"magersari", "Magersari Kota Mojokerto", -7.4800, 112.4445, "Magersari", "mjk-kota"),
    (r"kranggan", "Kranggan Kota Mojokerto", -7.4720, 112.4250, "Mojokerto Kota", "mjk-kota"),
    (r"beji|pasuruan", "Beji Pasuruan", -7.6580, 112.7280, "Beji", "mjk-kota"),
    (r"jetis", "Jetis Pusat", -7.4565, 112.4445, "Jetis", "mjk-kota"),
    (r"mojoanyar", "Mojoanyar Pusat", -7.4815, 112.4685, "Mojoanyar", "mjk-kota"),
]

ROUTES = {
    "mjk-kota": {"name": "Mojokerto Kota", "color": "#2563eb"},
    "trowulan": {"name": "Trowulan", "color": "#10b981"},
    "mojoagung": {"name": "Mojoagung – Jombang", "color": "#f59e0b"},
    "mojosari": {"name": "Mojosari – Pungging", "color": "#ef4444"},
    "puri": {"name": "Puri – Plososari", "color": "#8b5cf6"},
}

def geocode(record):
    text = ' '.join(filter(None, [
        record.get('9. Titik jemput yang diusulkan'),
        record.get('10. Patokan lokasi titik jemput (misal: Depan Indomaret X, Sebelah Masjid Y)'),
        record.get('13. Rute yang dibutuhkan'),
        record.get('7. Kecamatan domisili'),
        record.get('8. Kabupaten/Kota domisili'),
    ])).lower()
    for pat, label, lat, lng, area, rid in LANDMARKS:
        if re.search(pat, text, re.IGNORECASE):
            return label, lat, lng, area, rid
    return "Lokasi belum terverifikasi", -7.5, 112.45, "Lainnya", "mjk-kota"

responses = []
points_agg = {}
for d in data:
    label, lat, lng, area, rid = geocode(d)
    rec = {
        'n': (d.get('1. Nama lengkap') or '').strip(),
        'u': (d.get('3. Unit kerja') or '').strip(),
        's': (d.get('4. Status pegawai') or '').strip(),
        'k': (d.get('7. Kecamatan domisili') or '').strip().title(),
        'kb': (d.get('8. Kabupaten/Kota domisili') or '').strip().title(),
        'tj': (d.get('9. Titik jemput yang diusulkan') or '').strip(),
        'pt': (d.get('10. Patokan lokasi titik jemput (misal: Depan Indomaret X, Sebelah Masjid Y)') or '').strip(),
        'jr': (d.get('12. Jarak dari domisili ke titik jemput') or '').strip(),
        'rt': (d.get('13. Rute yang dibutuhkan') or '').strip(),
        'br': (d.get('14. Apakah bersedia bergabung dengan titik jemput terdekat?') or '').strip(),
        'jm': (d.get('17. Jam masuk kerja dominan') or '').strip(),
        'jp': (d.get('18. Jam pulang kerja dominan') or '').strip(),
        'pl': (d.get('19. Pola kerja') or '').strip(),
        'gn': (d.get('21. Apakah akan menggunakan layanan antar jemput jika rute tersedia?') or '').strip(),
        'fq': (d.get('22. Frekuensi penggunaan') or '').strip(),
        'al': (d.get('23. Alasan membutuhkan layanan antar jemput') or '').strip(),
        'lat': lat, 'lng': lng, 'pt_label': label, 'area': area, 'route': rid,
    }
    responses.append(rec)
    if label not in points_agg:
        points_agg[label] = {'label': label, 'lat': lat, 'lng': lng, 'area': area, 'route': rid, 'count': 0, 'members': []}
    points_agg[label]['count'] += 1
    points_agg[label]['members'].append(rec['n'])

points = sorted(points_agg.values(), key=lambda x: -x['count'])

print('// Auto-generated dataset from survey responses (110 entries)')
print('const HOSPITAL = {name:"RS Adhyaksa Jawa Timur", lat:-7.4995, lng:112.4548, address:"Jl. Raya Brangkal, Sooko, Kab. Mojokerto"};')
print('const ROUTES = ' + json.dumps(ROUTES, ensure_ascii=False) + ';')
print('const POINTS = ' + json.dumps(points, ensure_ascii=False) + ';')
print('const RESPONSES = ' + json.dumps(responses, ensure_ascii=False) + ';')
